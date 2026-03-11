import pandas as pd
from openpyxl.styles import PatternFill
import datetime


def generate_excel_report(comparison_results):
    """
    Generates an Excel report with:
    - Compliance score
    - Vendor match
    - Keywords
    - Audit questions
    - Explanation
    - Top 3 matches
    """

    data = []

    for result in comparison_results:

        bank_num = result.get("bank_section_number")
        vendor_section = result.get("matched_vendor_section")
        status = result.get("compliance_status")
        score = result.get("compliance_percentage")

        keywords = ", ".join(result.get("keywords", []))
        questions = " | ".join(result.get("audit_questions", []))
        explanation = result.get("explanation")

        top_matches = result.get("top_matches", [])

        row = {
            "Bank Section": bank_num,
            "Matched Vendor Section": vendor_section,
            "Compliance Status": status,
            "Compliance Score (%)": score,
            "Keywords": keywords,
            "Audit Questions": questions,
            "Explanation": explanation
        }

        # Add Top 3 Matches
        for i, match in enumerate(top_matches, start=1):
            row[f"Top {i} Vendor Section"] = match["vendor_para_number"]
            row[f"Top {i} Similarity Score"] = round(match["similarity_score"], 3)

        data.append(row)

    df = pd.DataFrame(data)

    # Generate timestamp filename
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"compliance_report_{timestamp}.xlsx"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:

        df.to_excel(writer, index=False, sheet_name="Compliance Report")

        ws = writer.sheets["Compliance Report"]

        # Apply color coding
        for row_idx, status in enumerate(df["Compliance Status"], start=2):

            if status == "Fully Compliant":
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            elif status == "Partially Compliant":
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            else:
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            ws[f"C{row_idx}"].fill = fill  # Compliance Status column

    return filename
